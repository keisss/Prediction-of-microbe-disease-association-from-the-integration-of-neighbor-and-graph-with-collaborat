import scipy.io as sio
import numpy as np
from openpyxl import Workbook

import fields


def load_diseases_to_array():
    f = open(fields.diseases_txt)
    d_list = []
    while True:
        line = f.readline().splitlines()
        if not line:
            break
        d_list.append(line[0])
    return d_list


def load_microbe_to_array():
    f = open(fields.microbes_txt)
    m_list = []
    while True:
        line = f.readline().splitlines()
        if not line:
            break
        m_list.append(line[0])
    return m_list


def load_mat_array():
    res_mat = sio.loadmat(fields.result_mat)
    return np.array(res_mat['S'])


if __name__ == '__main__':
    diseases = load_diseases_to_array()
    microbes = load_microbe_to_array()
    display_mat = dict()
    mat = load_mat_array()
    b = mat.size
    c = b
    for i in range(0, mat.shape[0]-1):
        for j in range(0, mat.shape[1]-1):
            name_i = diseases[i]
            name_j = microbes[j]
            # if name_i not in display_mat:
            #     display_mat[name_i] = dict()
            display_mat[(name_i, name_j)] = mat[i][j]
    sorted_dict = sorted(display_mat.items(), key=lambda item: item[1], reverse=True)

    wb = Workbook()  # 创建一个工作簿

    ws = wb.active  # 获取工作的激活工作表

    ws.cell(column=1, row=1, value='Disease')
    ws.cell(column=2, row=1, value='Microbe')
    ws.cell(column=3, row=1, value='Sim')
    for row in range(2, len(sorted_dict)):
        item = sorted_dict[row]
        ws.cell(column=1, row=row, value=item[0][0])
        ws.cell(column=2, row=row, value=item[0][1])
        ws.cell(column=3, row=row, value=item[1])
    wb.save('res.xls')
